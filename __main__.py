# -*-coding:Utf-8 -*
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.firefox.webdriver import WebDriver
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.firefox import GeckoDriverManager

driver: WebDriver = webdriver.Firefox(service=FirefoxService(GeckoDriverManager().install()))
wait: object = WebDriverWait(driver, 10)


def navigate_to_website(url: str):
    """
    Navigate to the website.
    Check that all job offers are displayed on the page.
    """
    driver.get(url)
    wait.until(ec.element_to_be_clickable(driver.find_element(By.CLASS_NAME, 'results-context-header')))
    jobs_offers_number: str = driver.find_element(By.CLASS_NAME, "results-context-header__job-count").text
    list_elements = []
    while len(list_elements) != int(jobs_offers_number):
        driver.execute_script("window.scrollTo(0,document.body.scrollHeight);")
        list_elements = driver.find_elements(By.XPATH, '//ul[@class="jobs-search__results-list"]/li')


def close_website():
    """
    Close the browser
    """
    driver.quit()


def get_job_datas():
    """
    Scrape datas (job position, company and offer link) from the website.
    Create an Excel file with the datas.
    """
    links = driver.find_elements(By.XPATH, "//li/div/a")
    positions = driver.find_elements(By.CLASS_NAME, "base-search-card__title")
    companies = driver.find_elements(By.XPATH, "//h4/a")
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Position"
    ws['B1'] = "Company"
    ws['C1'] = "Link"
    for index, position in enumerate(positions, start=2):
        ws['A' + str(index)] = position.text
    for index, company in enumerate(companies, start=2):
        ws['B' + str(index)] = company.text
    for index, link in enumerate(links, start=2):
        ws['C' + str(index)] = link.get_attribute("href")
    wb.save("rpa_jobs.xlsx")


if __name__ == '__main__':
    try:
        navigate_to_website("https://www.linkedin.com/jobs/search?keywords=RPA%20Developer&location=Portugal&geoId"
                            "=100364837&trk=public_jobs_jobs-search-bar_search-submit&position=1&pageNum=0")
        get_job_datas()
    except Exception as err:
        print(f"Unexpected {err=}, {type(err)=}")
        raise
    finally:
        close_website()
